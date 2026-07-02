from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook

from .models import SapCommand

REQUIRED_COLUMNS = ["tcode", "field_id", "value", "action"]


def _normalize(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_commands(excel_path: Path, sheet_name: str | None = None) -> List[SapCommand]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    header_row = [
        _normalize(cell.value).lower() if _normalize(cell.value) else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]

    missing = [column for column in REQUIRED_COLUMNS if column not in header_row]
    if missing:
        raise ValueError(
            "Excel template is missing required columns: " + ", ".join(missing)
        )

    idx = {name: header_row.index(name) for name in REQUIRED_COLUMNS}

    commands: list[SapCommand] = []
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        tcode = _normalize(row[idx["tcode"]].value)
        field_id = _normalize(row[idx["field_id"]].value)
        value = _normalize(row[idx["value"]].value)
        action = _normalize(row[idx["action"]].value)

        if not any([tcode, field_id, value, action]):
            continue

        commands.append(
            SapCommand(
                row_number=row_idx,
                tcode=tcode,
                field_id=field_id,
                value=value,
                action=action,
            )
        )

    return commands
